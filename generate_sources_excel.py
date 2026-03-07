"""Generate an Excel file listing all RSS sources for the veille IA."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sources RSS Veille IA"

# Styles
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
cat_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
cat_font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Headers
headers = ["Categorie", "Nom du site", "Lien du site", "Flux RSS", "Pertinence pour la veille"]
col_widths = [22, 28, 45, 55, 65]

for col, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = width

ws.row_dimensions[1].height = 25

# Data organized by category
sources = [
    # ── Presse Tech & Breaking News IA ──
    ("Presse Tech & Breaking News", [
        ("TechCrunch - AI",
         "https://techcrunch.com/category/artificial-intelligence/",
         "https://techcrunch.com/category/artificial-intelligence/feed/",
         "Reference mondiale pour les annonces startups et levees de fonds IA. Premier a couvrir les tours de financement (OpenAI, Anthropic, Mistral). Indispensable pour le suivi des mouvements strategiques."),
        ("The Verge AI",
         "https://www.theverge.com/ai-artificial-intelligence",
         "https://www.theverge.com/rss/ai-artificial-intelligence/index.xml",
         "Breaking news IA avec angle editorial fort. Couvre les controverses (Anthropic/Pentagone, vie privee), les lancements de modeles (GPT-5.4) et les enjeux geopolitiques IA. Source cle pour les signaux strategiques."),
        ("Ars Technica AI",
         "https://arstechnica.com/ai/",
         "https://arstechnica.com/ai/feed/",
         "Analyses techniques approfondies sur les modeles et produits IA. Couvre les proces (Gemini), la regulation (Californie/xAI), les controverses ethiques. Bon complement aux sources plus grand public."),
        ("Wired AI",
         "https://www.wired.com/tag/ai/",
         "https://www.wired.com/feed/tag/ai/latest/rss",
         "Enquetes et analyses geopolitiques IA (usage militaire, Iran, ByteDance). Couvre l'intersection IA/societe avec un angle investigatif rare dans les autres sources tech."),
        ("VentureBeat - AI",
         "https://venturebeat.com/category/ai/",
         "https://venturebeat.com/category/ai/feed/",
         "Couverture enterprise-first de l'IA : deploiements en entreprise, outils enterprise (Copilot, Einstein), strategies des DSI. Angle directement transposable aux grandes entreprises."),
        ("MIT Technology Review",
         "https://www.technologyreview.com/",
         "https://www.technologyreview.com/feed/",
         "Analyses de fond sur les tendances IA a 12-36 mois. Reference pour les etudes d'impact, les rapports chiffres et la prospective technologique. Credibilite academique."),
    ]),

    # ── Grands acteurs IA ──
    ("Grands acteurs IA (blogs officiels)", [
        ("Google AI Blog",
         "https://blog.google/technology/ai/",
         "https://blog.google/technology/ai/rss/",
         "Annonces officielles Google/DeepMind : lancements Gemini, percees en recherche, nouveaux produits. Source primaire pour suivre la strategie IA de Google."),
        ("Microsoft AI Blog",
         "https://blogs.microsoft.com/ai/",
         "https://blogs.microsoft.com/ai/feed/",
         "Strategie IA de Microsoft : Copilot, Azure AI, partenariat OpenAI. Impact direct sur les entreprises utilisant l'ecosysteme Microsoft (majorite du CAC 40)."),
        ("Microsoft France",
         "https://news.microsoft.com/fr-fr/",
         "https://news.microsoft.com/fr-fr/feed/",
         "Annonces specifiques au marche francais : partenariats locaux, deploiements Copilot en France, evenements. Contexte francais direct."),
        ("AWS Machine Learning Blog",
         "https://aws.amazon.com/blogs/machine-learning/",
         "https://aws.amazon.com/blogs/machine-learning/feed/",
         "Cas d'usage enterprise detailles sur AWS (Bedrock, SageMaker). Tutoriels et architectures de reference transposables. Pertinent pour les equipes cloud."),
        ("NVIDIA Newsroom",
         "https://nvidianews.nvidia.com/",
         "https://nvidianews.nvidia.com/rss.xml",
         "Annonces GPU/infrastructure IA : nouvelles puces, partenariats industriels, benchmarks. Impacte directement les choix d'infrastructure des Lab IA."),
        ("Meta AI",
         "https://ai.meta.com/",
         "https://rss.app/feeds/bdaGMyWkTwKKsATN.xml",
         "Recherche et modeles open source de Meta (LLaMA, etc.). Important pour la strategie open source IA et les alternatives aux modeles proprietaires."),
        ("Hugging Face Blog",
         "https://huggingface.co/blog",
         "https://huggingface.co/blog/feed.xml",
         "Ecosysteme open source IA : nouveaux modeles, benchmarks, outils. Hub de reference pour l'evaluation et le deploiement de modeles. Acteur francais majeur."),
    ]),

    # ── Reglementaire & Ethique ──
    ("Reglementaire & Ethique", [
        ("CNIL - Actualites",
         "https://www.cnil.fr/",
         "https://www.cnil.fr/fr/rss.xml",
         "Regulateur francais des donnees personnelles. Recommandations IA, sanctions, guides de conformite RGPD/IA. Obligatoire pour toute veille reglementaire IA en France."),
        ("LINC (CNIL)",
         "https://linc.cnil.fr/",
         "https://linc.cnil.fr/rss.xml",
         "Laboratoire d'innovation de la CNIL. Prospective sur les usages IA, vie privee, ethique. Signaux faibles reglementaires avant qu'ils deviennent des obligations."),
        ("Stanford HAI",
         "https://hai.stanford.edu/",
         "https://rss.app/feeds/Uq5zsW6lZLtmzBAs.xml",
         "Institut de reference mondiale sur l'IA centree sur l'humain. Rapports annuels AI Index, etudes d'impact, recommandations politiques. Credibilite academique de premier plan."),
        ("OECD AI Policy Observatory",
         "https://oecd.ai/",
         "https://rss.app/feeds/w96fbMAi1jrm945u.xml",
         "Politiques IA des pays OCDE : regulations, principes, comparaisons internationales. Essentiel pour anticiper les evolutions reglementaires au-dela de l'Europe."),
        ("World Economic Forum",
         "https://www.weforum.org/",
         "https://rss.app/feeds/Si39ht64GROA17C7.xml",
         "Rapports sur l'impact global de l'IA : emploi, competitivite, gouvernance. Cite par les dirigeants. Rapports Future of Jobs notamment."),
    ]),

    # ── RH & Impacts Sociaux ──
    ("RH & Impacts Sociaux", [
        ("France Strategie",
         "https://www.strategie.gouv.fr/",
         "https://www.strategie.gouv.fr/rss.xml",
         "Institution rattachee au Premier ministre. Etudes sur l'impact IA sur l'emploi en France, prospective metiers, dialogue social. Reference pour le contexte social francais."),
        ("Harvard Business Review",
         "https://hbr.org/",
         "https://rss.app/feeds/5keS8T8iWcv7B0qz.xml",
         "Articles management sur l'adoption IA : conduite du changement, leadership, transformation des metiers. Lu par les dirigeants. Angle managerial complementaire."),
    ]),

    # ── Cas d'usage & Marche ──
    ("Cas d'usage & Marche", [
        ("Journal du Net",
         "https://www.journaldunet.com/",
         "https://www.journaldunet.com/rss/",
         "Premiere source francophone en volume. Couvre l'IA en entreprise en France : cas d'usage, tendances marche, interviews DSI. Contexte 100% francais."),
        ("L'Usine Digitale",
         "https://www.usine-digitale.fr/",
         "https://rss.app/feeds/MfzXXWYYW25v1YMe.xml",
         "Transformation numerique de l'industrie francaise. Deploiements IA dans l'industrie, l'energie, les transports. Tres pertinent pour les entreprises industrielles du CAC 40."),
        ("McKinsey Insights",
         "https://www.mckinsey.com/featured-insights",
         "https://www.mckinsey.com/insights/rss",
         "Rapports de reference sur la maturite IA des grandes organisations. Etudes chiffrees (ROI, adoption, productivite). Cite dans les comites de direction."),
        ("CB Insights",
         "https://www.cbinsights.com/research/",
         "https://www.cbinsights.com/research/feed/",
         "Analyse marche IA : tendances investissement, cartographie des acteurs, rapports sectoriels. Utile pour le positionnement strategique et le benchmark."),
        ("Hub France IA",
         "https://www.hub-franceia.fr/",
         "https://www.hub-franceia.fr/feed/",
         "Association de reference de l'ecosysteme IA francais. Evenements, groupes de travail, publications sur l'IA en France. Reseau de l'ecosysteme national."),
        ("Maddyness",
         "https://www.maddyness.com/",
         "https://www.maddyness.com/feed/",
         "Ecosysteme startups et innovation en France. Couverture des startups IA francaises, levees de fonds, tendances. Angle entrepreneurial francais."),
        ("Capgemini Research Institute",
         "https://www.capgemini.com/insights/research-institute/",
         "https://rss.app/feeds/DTBeUkhTnXS7TtQH.xml",
         "Etudes enterprise sur l'IA : cas d'usage sectoriels, maturite IA, ROI. Acteur francais du conseil, etudes directement transposables aux grandes entreprises."),
    ]),

    # ── Cabinets de conseil ──
    ("Cabinets de conseil", [
        ("BCG Publications",
         "https://www.bcg.com/publications",
         "https://rss.app/feeds/OC8z0lI9kzPOGCY0.xml",
         "Etudes strategie et IA : transformation, cas clients, tendances sectorielles. Reference pour les presentations COMEX."),
        ("Deloitte Insights",
         "https://www2.deloitte.com/insights",
         "https://rss.app/feeds/Ry8cPt05z0pGDcz5.xml",
         "Rapports sur l'adoption IA en entreprise, tendances technologiques, State of AI. Approche audit/risque complementaire."),
        ("EY Newsroom",
         "https://www.ey.com/",
         "https://rss.app/feeds/C5JTkle1OFVzki8C.xml",
         "Actualites et etudes IA d'EY : conformite, audit, transformation. Angle reglementaire et risques."),
        ("KPMG Insights",
         "https://kpmg.com/",
         "https://rss.app/feeds/Hdh9fkcidmuErDsy.xml",
         "Etudes IA et data : gouvernance, confiance, adoption. Fort sur les enjeux de conformite et d'audit IA."),
        ("Accenture Newsroom",
         "https://newsroom.accenture.com/",
         "https://rss.app/feeds/aNsUrJodDVyXogQC.xml",
         "Annonces et etudes IA d'Accenture : cas clients, benchmarks d'adoption, tendances technologiques. Fort volume de deploiements IA."),
        ("Bain Insights",
         "https://www.bain.com/insights/",
         "https://rss.app/feeds/eVy2BnY5GsfOGVoC.xml",
         "Etudes strategie IA : ROI, modeles operationnels, transformation. Complement aux autres cabinets."),
        ("PwC Insights",
         "https://www.pwc.com/",
         "https://rss.app/feeds/Q1jvNVWtqI1E5TMA.xml",
         "Rapports Global AI Study, impact economique IA, enjeux de confiance. Reference pour les chiffres macro."),
        ("Roland Berger",
         "https://www.rolandberger.com/",
         "https://rss.app/feeds/9Z3q3nfx7PhpuUhE.xml",
         "Cabinet europeen de strategie. Etudes IA avec forte coloration industrie europeenne. Pertinent pour le contexte EU."),
    ]),

    # ── Presse Business ──
    ("Presse Business", [
        ("Business Insider",
         "https://www.businessinsider.com/",
         "https://www.businessinsider.com/rss",
         "Actualite business et tech grand public. Couvre les mouvements d'entreprise, les tendances management IA, les analyses marche. Angle business accessible."),
        ("Forbes",
         "https://www.forbes.com/",
         "https://rss.app/feeds/BJ580weB9yIHzJ6k.xml",
         "Listes, classements et analyses business IA. Interviews de dirigeants, tendances enterprise. Lu par les decideurs."),
    ]),

    # ── Presse Tech specialisee ──
    ("Presse Tech specialisee", [
        ("AIthority",
         "https://aithority.com/",
         "https://aithority.com/feed/",
         "Actualite IA enterprise : annonces produits, partenariats, analyses marche. Fort volume, couverture large de l'ecosysteme IA B2B."),
        ("CIO.com",
         "https://www.cio.com/",
         "https://www.cio.com/feed/",
         "Publication pour les DSI : strategies IT, deploiements IA, gouvernance. Directement adresse aux decideurs IT des grandes entreprises."),
        ("Numeum",
         "https://numeum.fr/",
         "https://numeum.fr/feed/",
         "Syndicat professionnel du numerique en France. Etudes marche, evenements, positions sur la regulation. Voix de l'industrie numerique francaise."),
    ]),

    # ── Agregateurs & Podcasts ──
    ("Agregateurs & Podcasts", [
        ("AI News Podcast",
         "https://media.rss.com/ai-news-chatgpt-openai-anthropic-claude/",
         "https://media.rss.com/ai-news-chatgpt-openai-anthropic-claude/feed.xml",
         "Podcast quotidien sur l'actualite IA. Resume des annonces cles. Utile comme filet de securite pour ne rater aucune news majeure."),
        ("AI News Digest",
         "",
         "https://rss.app/feeds/NUgCLd9ofzTDf58Q.xml",
         "Agregateur d'actualites IA. Couverture large des annonces et tendances. Complement pour capter les articles qui echappent aux autres sources."),
        ("AI Tech News",
         "",
         "https://rss.app/feeds/ZbnpQ8GpW7tKcvr0.xml",
         "Agregateur technique IA. Focus sur les sorties de modeles, benchmarks et outils. Bon pour le suivi technique."),
    ]),
]

row = 2
for category, feeds in sources:
    # Category header row
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.fill = cat_fill
        cell.border = thin_border
    ws.cell(row=row, column=1, value=category).font = cat_font
    ws.row_dimensions[row].height = 22
    row += 1

    for name, site_url, rss_url, description in feeds:
        ws.cell(row=row, column=1, value=category).alignment = wrap
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=name).alignment = wrap
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3, value=site_url).alignment = wrap
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).font = Font(color="0563C1", underline="single")
        ws.cell(row=row, column=4, value=rss_url).alignment = wrap
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).font = Font(color="0563C1", underline="single", size=9)
        ws.cell(row=row, column=5, value=description).alignment = wrap
        ws.cell(row=row, column=5).border = thin_border
        ws.row_dimensions[row].height = 50
        row += 1

# Freeze header
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:E{row - 1}"

output_path = r"C:\Users\pc\veille-ia\Sources_RSS_Veille_IA.xlsx"
wb.save(output_path)
print(f"Fichier cree: {output_path}")
print(f"{row - 2 - len(sources)} sources au total")
